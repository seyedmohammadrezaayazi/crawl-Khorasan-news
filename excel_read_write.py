import openpyxl
import os


class excel():

    def __init__(self, name, name_sheet):
        self.name = name + '.xlsx'

        PATH = self.name
        if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
            self.workbook = openpyxl.load_workbook(self.name)
        else:
            self.workbook = openpyxl.Workbook()
        self.worksheet=self.workbook.active
        if self.worksheet is None:
            self.workbook.create_sheet(index=1, title=name_sheet)
        self.max_row=self.worksheet.max_row
        self.workbook.save(self.name)


    def final(self):
        self.workbook.save(self.name)
        self.workbook.close()
